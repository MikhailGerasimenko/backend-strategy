from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path


@dataclass(frozen=True)
class Indicator:
    number: int
    name: str
    top_priority: bool


def load_indicators(path: Path) -> list[Indicator]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    indicators: list[Indicator] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]:
            continue
        number = int(row[0]) if row[0] is not None else len(indicators) + 1
        name = str(row[1]).strip()
        priority = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
        indicators.append(
            Indicator(
                number=number,
                name=name,
                top_priority=priority == "+",
            )
        )
    return indicators


def format_indicators_for_prompt(indicators: list[Indicator]) -> str:
    lines: list[str] = []
    for item in indicators:
        mark = " [топ-приоритет]" if item.top_priority else ""
        lines.append(f"{item.number}. {item.name}{mark}")
    return "\n".join(lines)


def load_format_excerpt(path: Path, max_chars: int = 7000) -> str:
    from pypdf import PdfReader

    text_parts: list[str] = []
    for page in PdfReader(str(path)).pages:
        text_parts.append(page.extract_text() or "")
    text = "\n".join(text_parts)
    text = " ".join(text.split())
    if len(text) <= max_chars:
        return text
    return text[: max_chars - 1] + "…"


def default_indicators_path(project_dir: Path) -> Path:
    return project_dir / "Список показателей.xlsx"


def default_format_pdf_path(project_dir: Path) -> Path:
    return project_dir / "news2026-03-19.pdf"
