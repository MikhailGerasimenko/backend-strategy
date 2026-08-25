from __future__ import annotations

import csv
import json
from collections import defaultdict
from pathlib import Path
from typing import Iterable

from .models import NewsItem, ParserHealth


NEWS_COLUMNS = [
    "source",
    "category",
    "title",
    "date",
    "url",
    "summary",
    "content",
    "language",
    "fetched_at",
    "status",
    "error",
    "relevance_match",
    "keyword_block",
    "keyword_match",
]
HEALTH_COLUMNS = ["source", "status", "items", "errors", "message", "fetched_at"]


def write_jsonl(path: Path, rows: Iterable[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as file:
        for row in rows:
            file.write(json.dumps(row, ensure_ascii=False) + "\n")


def write_csv(path: Path, rows: list[dict], columns: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_excel(path: Path, items: list[NewsItem], health: list[ParserHealth]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError:
        marker = path.with_suffix(".excel_skipped.txt")
        marker.write_text("Install openpyxl to enable Excel output.\n", encoding="utf-8")
        return

    rows_by_source: dict[str, list[dict]] = defaultdict(list)
    for item in items:
        rows_by_source[item.source].append(item.to_dict())

    wb = Workbook()
    wb.remove(wb.active)
    append_sheet(wb, "health", [row.to_dict() for row in health], HEALTH_COLUMNS)
    append_sheet(wb, "all_news", [item.to_dict() for item in items], NEWS_COLUMNS)
    for source_name, source_rows in rows_by_source.items():
        append_sheet(wb, safe_sheet_name(source_name), source_rows, NEWS_COLUMNS)
    wb.save(path)

    adjust_excel_widths(path, load_workbook)


def append_sheet(wb, sheet_name: str, rows: list[dict], columns: list[str]) -> None:
    sheet = wb.create_sheet(sheet_name)
    sheet.append(columns)
    for row in rows:
        sheet.append([row.get(column, "") for column in columns])


def safe_sheet_name(name: str) -> str:
    unsafe = "[]:*?/\\"
    cleaned = "".join("_" if char in unsafe else char for char in name)
    return (cleaned or "source")[:31]


def adjust_excel_widths(path: Path, load_workbook) -> None:
    wb = load_workbook(path)
    for sheet in wb.worksheets:
        for column_cells in sheet.columns:
            column_letter = column_cells[0].column_letter
            max_length = max((len(str(cell.value)) for cell in column_cells if cell.value), default=0)
            sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 80)
    wb.save(path)
